from config import *
from parsing import *
import openpyxl
import matplotlib.pyplot as plt
from matplotlib.offsetbox import (OffsetImage, AnnotationBbox)



def get_img(datte, clas: str):
  wookbook = openpyxl.load_workbook(f"./outputs/{datte} schedule.xlsx")
  ws = wookbook.active
  dict_classes = {
    '5': ['C 2', 'E 2'],
    '6': ['G 2', 'I 2'],
    '7': ['K 2', 'M 2'],
    '8': ['C 11', 'E 11'],
    '9': ['G 11', 'I 11'],
    '10': ['K 12'],
    '11': ['M 12']
  }
  time_classes = 'B 2'.split(' ')
  
  if clas != '10' and clas != '11':
    a_class = dict_classes[clas][0].split(' ')
    b_class = dict_classes[clas][1].split(' ')
    
    times = []
    a = []
    b = []

    for i in range(int(a_class[1]), int(a_class[1])+9):
      a.append(ws[f'{a_class[0]}{i}'].value)
      b.append(ws[f'{b_class[0]}{i}'].value)
      times.append(ws[f'{time_classes[0]}{i}'].value)
    
    #define figure and axes
    fig, ax = plt.subplots()
    im = plt.imread('./bb.png')
    imagebox = OffsetImage(im, zoom = 0.05)
    ab = AnnotationBbox(imagebox, (0.05, 1), frameon = False)
    ax.add_artist(ab)
    #create values for table
    table_data=[]
    for g in zip(a,b, times):
      table_data.append(tuple([g[2], g[0], g[1]]))

    #create table
    plt.rcParams['font.family'] = 'Helvetica'
    table = ax.table(cellText=table_data, loc='center')
    csfont = {'fontname':'Helvetica'}
    # plt.title('f.png')
    ax.set_title(f"{clas} класс на {datte}", x=0.25, y=0.96, fontsize=10, fontweight='bold', **csfont, color='#048B7B')
    
    table[(0, 0)].set_facecolor("#048B7B")
    table[(0, 1)].set_facecolor("#048B7B")
    table[(0, 2)].set_facecolor("#048B7B")
    table[(0,0)].set_text_props(weight='bold', color='w')
    table[(0,1)].set_text_props(weight='bold', color='w')
    table[(0,2)].set_text_props(weight='bold', color="w")
    ax.axis('off')
    table.scale(1,2)
    #display table
    plt.savefig(f'./outputs/{datte} {clas}.png', dpi=800)
  else:
    time_classes = 'B 2'.split(' ')
    sep = dict_classes[clas][0].split(' ')
    
    times = []
    res = []

    for i in range(int(sep[1]), int(sep[1])+9):
      res.append(ws[f'{sep[0]}{i}'].value)
      times.append(ws[f'{time_classes[0]}{i}'].value)
    #define figure and axes
    fig, ax = plt.subplots()
    im = plt.imread('./bb.png')
    imagebox = OffsetImage(im, zoom = 0.05)
    ab = AnnotationBbox(imagebox, (0.04, 1), frameon = False)
    ax.add_artist(ab)
    #create values for table
    table_data=[]
    for g in zip(res, times):
      table_data.append(tuple([g[1], g[0]]))



    #create table
    plt.rcParams['font.family'] = 'Helvetica'
    csfont = {'fontname':'Helvetica'}
    table = ax.table(cellText=table_data, loc='center')
    ax.set_title(f"{clas} класс на {datte}", x=0.26, y=0.96, fontsize=10, fontweight='bold', **csfont, color='#048B7B')
    #modify table
    table[(0, 0)].set_facecolor("#048B7B")
    table[(0, 1)].set_facecolor("#048B7B")
    table[(0,0)].set_text_props(weight='bold', color='w')
    table[(0,1)].set_text_props(weight='bold', color='w')
    table.scale(1,2)
    ax.axis('off')

    #display table
    plt.savefig(f'./outputs/{datte} {clas}.png', dpi=800)


def get_hw():
  ls = parsing_homework(get_data(0))[0]
  hws = parsing_homework(get_data(0))[1]
  ls.insert(0, str('Предметы'))
  hws.insert(0, str('Домашнее задание'))

  log = []
  for n in zip(hws, ls):
    if len(n[0]) > 100:
      log.append(n)
      # s = n.split(' ')
      # print(s)
      # s[int(len(s)/2)] = '\n'
      # s[int(len(s)/2/2)] = '\n'
      hws.insert(hws.index(n[0]), "".join(n[0].split(' ')[:4]))
      hws.remove(n[0])
  res = []
  for l in log:
    res.append(f"*{l[1]}*: {l[0]}\n")
  fig, ax = plt.subplots()
  im = plt.imread('./bb.png')
  imagebox = OffsetImage(im, zoom = 0.06)
  ab = AnnotationBbox(imagebox, (0.5, 1), frameon = False)
  ax.add_artist(ab)
  table_data=[]
  for g in zip(ls, hws):
    table_data.append(tuple([g[0], str(g[1])]))
  plt.rcParams['font.family'] = 'Helvetica'
  table = ax.table(cellText=table_data, loc='center')
  
  table[(0, 0)].set_facecolor("#048B7B")
  table[(0, 1)].set_facecolor("#048B7B")
  table[(0,0)].set_text_props(weight='bold', color='w')
  table[(0,1)].set_text_props(weight='bold', color='w')
  table.scale(1, 4)

  ax.axis('off')
  plt.savefig(f'./outputs/hw {tomorrow()}.png', dpi=800)
  return res


def FoodParser(date, timeToFood: str):
  foodXlsx = openpyxl.load_workbook(f"./outputs/{date} food.xlsx")
  excelCursor = foodXlsx.active
  resultForDiner = 0
  for find in range(3, 15):
    resultForDiner += 1
    if excelCursor[f'B{find}'].value == "Итог по набору":
      break
    

  elementsPosition = {
    'breakfast': ['B 3', 'D 3'],
    'dinner': [f'B {3+resultForDiner}', f'D {3+resultForDiner}']
  }
  typeOfFoodResults = []
  foodResults = []

  if timeToFood == "Завтрак":

    breakfastType = elementsPosition['breakfast'][0].split(' ')
    breakfastFood = elementsPosition['breakfast'][1].split(' ')

    for i in range(int(breakfastType[1]), int(breakfastType[1])+10):
      if excelCursor[f'{breakfastType[0]}{i}'].value == "Итог по набору":
        break
      typeOfFoodResults.append(excelCursor[f'{breakfastType[0]}{i}'].value)
      foodResults.append(excelCursor[f'{breakfastFood[0]}{i}'].value)
    fig, ax = plt.subplots()
    im = plt.imread('./bb.png')
    imagebox = OffsetImage(im, zoom = 0.05)
    ab = AnnotationBbox(imagebox, (0.05, 0.9), frameon = False)
    ax.add_artist(ab)
    #create values for table
    table_data=[]
    for g in zip(typeOfFoodResults, foodResults):
      table_data.append(tuple([g[0], g[1]]))

    plt.rcParams['font.family'] = 'Helvetica'
    table = ax.table(cellText=table_data, loc='center')
    csfont = {'fontname':'Helvetica'}
    # plt.title('f.png')
    ax.set_title(f"Что будем хавать {date}", x=0.32, y=0.86, fontsize=10, fontweight='bold', **csfont, color='#048B7B')
    
    table[(0, 0)].set_facecolor("#048B7B")
    table[(0, 1)].set_facecolor("#048B7B")
    table[(0,0)].set_text_props(weight='bold', color='w')
    table[(0,1)].set_text_props(weight='bold', color='w')
    ax.axis('off')
    table.scale(1,2)
    #display table
    plt.savefig(f'./outputs/{date} food.png', dpi=800)
  
  else:
    dinnerType = elementsPosition['dinner'][0].split(' ')
    dinnerFood = elementsPosition['dinner'][1].split(' ')
    for i in range(int(dinnerType[1]), int(dinnerType[1])+10):
      
      if excelCursor[f'{dinnerType[0]}{i}'].value == "Итог по набору":
        break
      typeOfFoodResults.append(excelCursor[f'{dinnerType[0]}{i}'].value)
      foodResults.append(excelCursor[f'{dinnerFood[0]}{i}'].value)
    

    fig, ax = plt.subplots()
    im = plt.imread('./bb.png')
    imagebox = OffsetImage(im, zoom = 0.05)
    ab = AnnotationBbox(imagebox, (0.05, 0.9), frameon = False)
    ax.add_artist(ab)
    #create values for table
    table_data=[]
    table_data.append(tuple(["Раздел", "Блюдо"]))
    for g in zip(typeOfFoodResults, foodResults):
      table_data.append(tuple([g[0], g[1]]))

    plt.rcParams['font.family'] = 'Helvetica'
    table = ax.table(cellText=table_data, loc='center')
    csfont = {'fontname':'Helvetica'}
    # plt.title('f.png')
    ax.set_title(f"Что будем хавать {date}", x=0.32, y=0.86, fontsize=10, fontweight='bold', **csfont, color='#048B7B')
    
    table[(0, 0)].set_facecolor("#048B7B")
    table[(0, 1)].set_facecolor("#048B7B")
    table[(0,0)].set_text_props(weight='bold', color='w')
    table[(0,1)].set_text_props(weight='bold', color='w')
    ax.axis('off')
    table.scale(1,2)
    #display table
    plt.savefig(f'./outputs/{date} food.png', dpi=800)
